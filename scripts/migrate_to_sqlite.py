import openpyxl
import sqlite3
import re
import os

excel_path = "data/APOTEK HF.xlsx"
db_path = "data/apotek.db"

# Ensure data directory exists
os.makedirs("data", exist_ok=True)

# Connect to SQLite
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

print("Loading Excel workbook (data_only=True)...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

def clean_col_name(name):
    if name is None:
        return "unnamed_col"
    name = str(name).strip()
    name = re.sub(r'[^a-zA-Z0-9_]', '_', name)
    name = re.sub(r'_+', '_', name)
    name = name.strip('_')
    return name.lower()

for sheet_name in wb.sheetnames:
    print(f"\nProcessing sheet: {sheet_name}")
    sheet = wb[sheet_name]
    
    # Get all rows
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        print(f"Sheet {sheet_name} is empty, skipping.")
        continue
        
    if not headers or all(h is None for h in headers):
        print(f"Sheet {sheet_name} has no valid headers, skipping.")
        continue
        
    # Clean headers and ensure they are unique
    clean_headers = []
    seen = set()
    for h in headers:
        c_name = clean_col_name(h)
        if not c_name:
            c_name = "unnamed"
        original_c_name = c_name
        counter = 1
        while c_name in seen:
            c_name = f"{original_c_name}_{counter}"
            counter += 1
        seen.add(c_name)
        clean_headers.append(c_name)
        
    print(f"Normalized Columns: {clean_headers}")
    
    # Read all data rows
    data_rows = []
    for r in rows_iter:
        # Check if the row has any non-None data
        if any(cell is not None for cell in r):
            # Trim row or pad to match clean_headers length
            padded_row = list(r[:len(clean_headers)])
            if len(padded_row) < len(clean_headers):
                padded_row += [None] * (len(clean_headers) - len(padded_row))
            # Convert non-standard objects to string (like array formulas if any remain)
            for idx, val in enumerate(padded_row):
                if val is not None and not isinstance(val, (int, float, str, bytes)):
                    padded_row[idx] = str(val)
            data_rows.append(tuple(padded_row))
            
    if not data_rows:
        print(f"No data rows found in {sheet_name}, skipping table creation.")
        continue
        
    # Drop existing table if any
    safe_table_name = clean_col_name(sheet_name)
    cursor.execute(f"DROP TABLE IF EXISTS {safe_table_name}")
    
    # Create Table
    col_defs = ", ".join([f'"{h}" TEXT' for h in clean_headers]) # default to TEXT for flexibility
    create_sql = f'CREATE TABLE "{safe_table_name}" ({col_defs})'
    cursor.execute(create_sql)
    
    # Insert Rows
    placeholders = ", ".join(["?"] * len(clean_headers))
    insert_sql = f'INSERT INTO "{safe_table_name}" VALUES ({placeholders})'
    
    # Execute batch insert
    cursor.executemany(insert_sql, data_rows)
    conn.commit()
    print(f"Created table '{safe_table_name}' and inserted {len(data_rows)} rows.")

conn.close()
print("\nMigration complete! SQLite database created at:", db_path)
